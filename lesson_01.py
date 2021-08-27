# str1="metoo"
#print(str1.find("t"))
#print(str1.index("o"))
#print(str1.count("o"))
#st2=str1.replace("t","s")
#print(st2)
#tt=str1.split(",")
#print(tt)

# name="scream"
# height="168cm"
# weight="50kg"
# 'int('''----{}是84期的学霸
# 姓名:{}
# 身高:{}
# 体重:{}
# '''.format(name,name,height,weight))

# print(2+3)
# # print("scre"+"am")
# # print("scre"+str(5))
# # print(9-5)
# # print("我爱你"*5)

# a=5
# a=a+5
# print(a)

# print(5<4)
# print("操作成功"=="操作成功")

# print(6>7 or 7>6)
# print( not 7>=8)

# str="scream"
# print("crazy"in str)
# print("crazy"not in str)

# list1 = [ 101 , 102 , 103 ,[1,2,3]]
# print(len(list1))
# print(list1)
# print(list1[2])
# print(list1[3])
# list1.append("春天")
# print(list1)
# list1.insert(3,"bat")
# print(list1)
# list1.extend(["夏天","秋天","冬天"])
# print(list1)
# list1.pop(1)
# print(list1)
# list1.remove(101)
# print(list1)
# # list1.clear()
# # print(list1)
# list1[1]="sky"
# print(list1)

# name=input("请输入姓名:")
# gender=input("请输入性别:")
# age=input("请输入年龄：")
# print("姓名:"+name)
# print("性别："+gender)
# print("年龄:"+age)

# str1 = 'python hello aaa 123123aabb'
# print(str1[0:6])
# print("o" in str1)
# print("a" in str1)
# print("he" in str1)
# print("ab" in str1)
# str2=str1.replace("python","lemon")
# print(str2)
# print(str1.find("aaa"))

# dict_info={"name":"Rella","age":18,"height":"165cm","weight":"45kg"}
# print(dict_info["name"])
# print(dict_info.get("name"))
# print(dict_info.values())
# print(dict_info.items())
# print(dict_info.keys())
#增加，修改
# dict_info["hobby"]="sing"#赋值新的键值对，key不存在，新增
# dict_info["age"]="19"#赋值新的键值对，key存在，修改
# print(dict_info)
# #增加多个--字典合并
# dict_info.update({"city":"北京","gender":"female"})
# print(dict_info)
# #删除
# dict_info.pop("weight")#无序-没有最后一个=只能指定删除
# print(dict_info)
# print(len(dict_info))
#
# set1={11,2,3,4,5,5,5,5}
# print(type(set1))
# print(set1)
# list1=[11,22,33,44,44,55,55,11]
# print(list1)
# set2=set(list1)
# print(set2)
# list2=list(set2)
# print(list2)

# money=int(input("请输入你的财产金额："))#input输入内容-字符串
# print(type(money))
# if money>=200:
#     print("买大房子!")
# elif money>=100:
#     print("付首付")
# elif money>=50:
#     print("买车")
# elif money>=10:
#     print("买台电脑，吃喝玩乐")
# else:
#     print("打工")

# count=0
# str3="柠檬班全程班的学生是最棒的"
# for i in str3:
#     if i=="全":
#         #break
#         continue
#     print(i)
#     print("*"*20)
#     count +=1
# print(count)
# print(len(str3))

# for i in range(5):
#     print(i)

a=[1,2,6,'summer']
# print("i"in"a")
# print("i"not in"a")
#
# num=int(input("请输入你的班级上课人数:"))
# if num>5:
#     print(num)
#
# list1 =  [{"name":"Rella","gender":"women","age":"16","city":"西安"},{"name":"纯纯","gender":"women","age":"17","city":"北京"},{"name":"二丫","gender":"women","age":"18","city":"杭州"},{"name":"小梨","gender":"women","age":"16","city":"天津"},{"name":"东东","gender":"men","age":"16","city":"临潼"},{"name":"熊熊","gender":"men","age":"16","city":"昆明"}]
# print(list1)

# print('66666')#无用代码
# def good_job(salary,bonus,subsidy=500,*args,**kwargs):
#     print("salary参数的值是:{}".format(salary))
#     print("bonus参数的值是:{}".format(bonus))
#     print("subsidy参数的值是:{}".format(subsidy))
#     print("args参数的值是:{}".format(args))
#     print("kwargs参数的值是:{}".format(kwargs))
#     sum1=salary+bonus+subsidy
#     for i in args:
#         sum1 += i
#     for j in kwargs:
#         sum1 += kwargs.get(j)
#     print("这个工作的工资总和是:{}".format(sum1))
#     return sum1
# result=good_job(10000,5000,800,100,200,300,400,aa=100,bb=200,cc=300)#函数的调用-参数的传入=实参
# print(result)
# #变量来接受函数的调用--返回值
# if result > 10000:
#     print("这是一个好的工作！")
# else:
#     print("这工作不行")

# dic1={"name":"scream","age":"19"}
# dict2=dict(name="scream",age=19)
# print(type(dict2))

import requests#导入库
import pprint
import jsonpath
#注册
# url_register="http://47.115.15.198:7001/smarthome/user/register"#定义变量接受 接口地址
# test_date={
#   "phone": "15335541763",
#   "pwd": "lemon123",
#   "rePwd":"lemon123",
#   "userName":"挤挤总会有的",
#   "verificationCode":"lemon"}#定义参数
# head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}#定义请求头部
# response = requests.post(url=url_register,json=test_date,headers=head)#发送接口请求--响应消息
# print(response.json())#得到json格式响应体
#
# #登录
# url_login = "http://47.115.15.198:7001/smarthome/user/login"
# login_date = {"pwd":"lemon123","userName":"挤挤总会有的"}
# head_login = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
# res_login = requests.post(url=url_login,json=login_date,headers=head_login)
# pprint.pprint(res_login.json())
#
# #登录完后取值--id token
# # id = res_login.json().get("data")
# # token = res_login.json().get("data").get("token_info").get("token")
# id = jsonpath.jsonpath(res_login.json(),"$..id")[0]
# token = jsonpath.jsonpath(res_login.json(),"$..token")[0]
# pprint.pprint(id)
# pprint.pprint(token)
#
# #完善信息接口
# url_complete="http://47.115.15.198:7001/smarthome/merchant/complete"
# date_complete={
#   "address":"陕西省雁塔区电子城街道锦业路",
#   "establishDate":"2021-04-02",
#   "legalPerson":"沈谋仁",
#   "licenseCode":"555443355522333",
#   "licenseUrl":"http://127.0.0.1/smarthome/aaa.jpg",
#   "merchantName":"陕西安远门科技有限公司",
#   "merchantType":2,
#   "registerAuthority":"雁塔区电子城派出所",
#   "tel":"15335541763",
#   "userId":id,
#   "validityDate":"2033-05-02"}
# head_complete={"X-Lemonban-Media-Type":"lemonban.v2",
# "Content-Type":"application/json",
#                "Authorization":"Bearer"+token}
# result=requests.put(url=url_complete,json=date_complete,headers=head_complete)
# pprint.pprint(result.json())

from openpyxl import load_workbook # 导入部分
import requests
#读取函数
def read_data(filename,sheetname):
  wb = load_workbook(filename) # 加载工作簿
  sheet = wb[sheetname] #表单
  # cell = sheet.cell(row=1,column=1) #单元格
  # print(cell.value) # 单元格的内容
  # cell.value = "测试用例" #写入单元格的内容
  # wb.save("testcase_api_wuye.xlsx") #保存文件，另存为
  # print(cell.value)
  max_row = sheet.max_row #获取最大行数
  cases = [] #空列表--所有测试用例
  for i in range(2,max_row+1):
    case = dict(
    case_id = sheet.cell(row=i,column=1).value,
    url = sheet.cell(row=2,column=5).value,
    data = sheet.cell(row=2,column=6).value,
    expected_result = sheet.cell(row=2,column=7).value)
    cases.append(case)
  return cases
test_case=read_data("testcase_api_wuye.xlsx","login")  #返回值

#发送接口请求函数
def api_request(url_wuye,test_data):
  head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}#定义请求头部
  response = requests.post(url=url_wuye,json=test_data,headers=head) #发送接口请求的 --响应消息
  return  response.json() #得到json 格式响应体

#回写结果函数
def write_date(filename,sheetname,row,column,final_result):
  wb = load_workbook(filename)
  sheet = wb[sheetname]
  sheet.cell(row,column).value = final_result
  wb.save(filename)

#执行测试
def execute_fun(filename,sheetname):
  cases = read_data(filename,sheetname)
  for case in cases:
    case_id = case.get("case_id")
    url = case["url"]
    data = case["data"]#字符串
    data = eval(data) #字符串强制转化为 字典
    expected_result = case["expected_result"]
    expected_result = eval(expected_result)
    expect_msg = expected_result["msg"]
    real_result = api_request(url_wuye=url, test_data=data)  # 调用接口请求函数
    real_msg = real_result["msg"]
    print("预期结果是：{}".format(expect_msg))
    print("执行结果是：{}".format(real_msg))
    if expect_msg == real_msg:
      print("第{}条测试用例通过！".format(case_id))
    else:
      print("第{}条测试结果不通过".format(case_id))
      final_result = "Falied"
      print("*"*20)
      write_date(filename,sheetname,case_id+1,8,final_result=final_result)#写入函数结果

execute_fun("testcase_api_wuye.xlsx","login")

